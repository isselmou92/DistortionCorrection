import nibabel as nib
import matplotlib.pyplot as plt
from matplotlib.colors import LinearSegmentedColormap
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# Define the colors for the colormap
colors = ["blue", "aquamarine", "yellow", "red"]

# Create the colormap
cmap = LinearSegmentedColormap.from_list("custom_bright_colormap", colors)

# Path to your NIfTI file
ge_deformation_field_shim1_path = 'data/deformable/displacement_fields_shim1_shim2/shim1/displacement_field_shim1_lps.nii'
ge_deformation_field_shim2_path = 'data/deformable/displacement_fields_shim1_shim2/shim2/displacement_field_shim2_lps.nii'

ge_deformation_field_shim1_b0_corrected_path = 'b0_no_b0_displacement_fields/Displacement_field_Phantom_MR_Shim1_B0_Deformable_B0_Corrected.nii'
ge_deformation_field_shim2_b0_corrected_path = 'b0_no_b0_displacement_fields/Displacement_field_Phantom_MR_Shim2_Deformable_B0_Corrected.nii'

# Load the NIfTI file
ge_deformation_field_nii_shim1 = nib.load(ge_deformation_field_shim1_path)
ge_deformation_field_nii_shim2 = nib.load(ge_deformation_field_shim2_path)

ge_deformation_field_shim1_b0_corrected = nib.load(ge_deformation_field_shim1_b0_corrected_path)
ge_deformation_field_shim2_b0_corrected = nib.load(ge_deformation_field_shim2_b0_corrected_path)

# Extract the data
ge_deformation_field_data_shim1 = ge_deformation_field_nii_shim1.get_fdata()
ge_deformation_field_data_shim2 = ge_deformation_field_nii_shim2.get_fdata()

ge_deformation_field_shim1_b0_corrected_data = ge_deformation_field_shim1_b0_corrected.get_fdata()
ge_deformation_field_shim2_b0_corrected_data = ge_deformation_field_shim2_b0_corrected.get_fdata()

# Remove the singleton dimension if necessary
ge_deformation_field_data_squeezed_shim1 = np.squeeze(ge_deformation_field_data_shim1)
ge_deformation_field_data_squeezed_shim2 = np.squeeze(ge_deformation_field_data_shim2)

ge_deformation_field_shim1_b0_corrected_data_squeezed = np.squeeze(ge_deformation_field_shim1_b0_corrected_data)
ge_deformation_field_shim2_b0_corrected_data_squeezed = np.squeeze(ge_deformation_field_shim2_b0_corrected_data)

# Calculate the magnitude of the displacement vectors
ge_magnitudes_shim1 = np.linalg.norm(ge_deformation_field_data_squeezed_shim1, axis=-1)
ge_magnitudes_shim2 = np.linalg.norm(ge_deformation_field_data_squeezed_shim2, axis=-1)

ge_magnitudes_shim1_b0_corrected = np.linalg.norm(ge_deformation_field_shim1_b0_corrected_data_squeezed, axis=-1)
ge_magnitudes_shim2_b0_corrected = np.linalg.norm(ge_deformation_field_shim2_b0_corrected_data_squeezed, axis=-1)

# Coronal projection (mean across Y-axis) and transpose
ge_coronal_projection_shim1 = np.rot90(np.mean(ge_magnitudes_shim1, axis=1), 2).T
ge_coronal_projection_shim2 = np.rot90(np.mean(ge_magnitudes_shim2, axis=1), 2).T

ge_coronal_projection_shim1_b0_corrected = np.rot90(np.mean(ge_magnitudes_shim1_b0_corrected, axis=1), 2).T
ge_coronal_projection_shim2_b0_corrected = np.rot90(np.mean(ge_magnitudes_shim2_b0_corrected, axis=1), 2).T

fig, axes = plt.subplots(1, 4, figsize=(16, 11), constrained_layout=True)

# Compute ranges dynamically for the two pairs of images
vmin1 = min(ge_coronal_projection_shim1.min(), ge_coronal_projection_shim2.min())
vmax1 = max(ge_coronal_projection_shim1.max(), ge_coronal_projection_shim2.max())

vmin2 = min(ge_coronal_projection_shim1_b0_corrected.min(), ge_coronal_projection_shim2_b0_corrected.min())
vmax2 = max(ge_coronal_projection_shim1_b0_corrected.max(), ge_coronal_projection_shim2_b0_corrected.max())

# Plot Shim-1
im1 = axes[0].imshow(ge_coronal_projection_shim1, cmap=cmap, vmin=vmin1, vmax=vmax1)
axes[0].set_title('GRE Shim-1', fontsize=14)
axes[0].axis('off')

# Plot Shim-2
im2 = axes[1].imshow(ge_coronal_projection_shim2, cmap=cmap, vmin=vmin1, vmax=vmax1)
axes[1].set_title('GRE Shim-2', fontsize=14)
axes[1].axis('off')

# Plot Shim-1 B0 Corrected
im3 = axes[2].imshow(ge_coronal_projection_shim1_b0_corrected, cmap=cmap, vmin=vmin2, vmax=vmax2)
axes[2].set_title('GRE Shim-1 B0 Corrected', fontsize=14)
axes[2].axis('off')

# Plot Shim-2 B0 Corrected
im4 = axes[3].imshow(ge_coronal_projection_shim2_b0_corrected, cmap=cmap, vmin=vmin2, vmax=vmax2)
axes[3].set_title('GRE Shim-2 B0 Corrected', fontsize=14)
axes[3].axis('off')

# Add colorbar for the first two plots
cbar1 = fig.colorbar(im1, ax=axes[:2], orientation='vertical', fraction=0.05, pad=0.04)
cbar1.set_label('Displacement Magnitude (Shim 1 & 2)', fontsize=12)

# Add colorbar for the last two plots
cbar2 = fig.colorbar(im3, ax=axes[2:], orientation='vertical', fraction=0.05, pad=0.04)
cbar2.set_label('Displacement Magnitude (B0 Corrected)', fontsize=12)

plt.show()


# Function to calculate statistics
def calculate_statistics(array):
    return {
        'Min': np.min(array),
        'Max': np.max(array),
        'Mean': np.mean(array),
        'Median': np.median(array),
        'Std Dev': np.std(array),
        'Variance': np.var(array)
    }


# Creating a summary table
summary_data = {
    'ge_coronal_projection_shim1': calculate_statistics(ge_coronal_projection_shim1),
    'ge_coronal_projection_shim2': calculate_statistics(ge_coronal_projection_shim2),
    'ge_coronal_projection_shim1_b0_corrected': calculate_statistics(ge_coronal_projection_shim1_b0_corrected),
    'ge_coronal_projection_shim2_b0_corrected': calculate_statistics(ge_coronal_projection_shim2_b0_corrected)
}

# Convert to DataFrame for better readability
summary_df = pd.DataFrame(summary_data).T

# Write the summary DataFrame to an Excel file
excel_file_path = 'b0_no_b0_displacement_fields/displacement_field_summary.xlsx'
summary_df.to_excel(excel_file_path, index=True, sheet_name="Summary")

# Load the workbook and apply formatting
workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook['Summary']

# Styling: Set bold font for header and index, center alignment, and light blue fill for header
header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
header_font = Font(bold=True)
center_alignment = Alignment(horizontal="center", vertical="center")

# Apply formatting to header row and index column
for cell in sheet["1:1"]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_alignment

for cell in sheet["A"]:
    cell.font = header_font

# Set column widths for better readability
for column_cells in sheet.columns:
    max_length = max(len(str(cell.value)) for cell in column_cells)
    column_letter = column_cells[0].column_letter
    sheet.column_dimensions[column_letter].width = max_length + 2

# Save the formatted workbook
workbook.save(excel_file_path)

print(f"Excel file saved as {excel_file_path}")
